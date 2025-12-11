import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import os

st.set_page_config(layout="wide", page_title="时间线分析工具")

st.title("📊 电梯时间线数据可视化")

# 侧边栏：文件上传
st.sidebar.header("1. 数据加载")
uploaded_file = st.sidebar.file_uploader("上传 Excel 文件", type=["xlsx"])

# 默认文件路径（方便测试）
DEFAULT_FILE_PATH = '/Users/seasong/Nutstore Files/我的坚果云/python/timeline_ai/data/23N4B16-474-43等_20251211084439_Both.xlsx'

def get_rgb_color(color_obj, wb):
    """
    辅助函数：尝试将 openpyxl 的颜色对象转换为 RGB 字符串 (#RRGGBB)。
    """
    if not color_obj:
        return None
    
    # 1. RGB 类型
    if color_obj.type == 'rgb':
        # 有时候 RGB 是 '00RRGGBB'，需要截取
        if len(color_obj.rgb) == 8:
            return '#' + color_obj.rgb[2:]
        return '#' + color_obj.rgb
    
    # 2. Theme 类型 (比较复杂，这里做简单近似或忽略)
    # 真正的 Theme 颜色转换需要解析 theme.xml，比较繁琐。
    # 这里为了简便，如果遇到 theme color，暂时返回 None 或默认值。
    # 也可以引入 wcag_contrast_ratio 等库来计算，但为了保持无依赖，先忽略。
    if color_obj.type == 'theme':
        # 尝试一些简单的映射，或者直接返回 None
        # print(f"Theme color found: {color_obj.theme}, tint: {color_obj.tint}")
        return None
        
    # 3. Indexed 类型
    if color_obj.type == 'indexed':
        # 标准 Excel 调色板
        # 可以硬编码一个 lookup table，但这里暂略
        return None
        
    return None

@st.cache_data
def load_data_with_styles(file):
    """
    加载 Excel 数据，并提取背景色、字体色和备注。
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    
    # 查找时间线 Sheet
    target_sheet_name = None
    # openpyxl 的属性是 sheetnames，不是 sheet_names
    for name in wb.sheetnames:
        if "时间线" in name or "Timeline" in name:
            target_sheet_name = name
            break
    
    if not target_sheet_name:
        return None, None, None, f"未找到名为 '时间线' 或 'Timeline' 的工作表。可用工作表: {wb.sheetnames}"
    
    ws = wb[target_sheet_name]
    
    data = []
    styles = [] # 存储 CSS 样式字符串
    comments = [] # 存储备注信息
    
    # 获取表头
    headers = [cell.value for cell in ws[1]]
    
    # 遍历数据行
    for row in ws.iter_rows(min_row=2):
        row_data = []
        row_style = []
        row_comment = []
        
        for cell in row:
            # 1. 值
            row_data.append(cell.value)
            
            # 2. 样式 (背景色 + 字体色)
            cell_css = ""
            
            # 背景色
            bg_color = get_rgb_color(cell.fill.fgColor, wb) if cell.fill else None
            # 如果背景色是白色或透明，通常忽略
            if bg_color and bg_color.upper() not in ['#000000', '#FFFFFF', '#00FFFFFF']: 
                 cell_css += f"background-color: {bg_color}; "
            
            # 字体色
            font_color = get_rgb_color(cell.font.color, wb) if cell.font else None
            if font_color and font_color.upper() not in ['#000000', '#FFFFFF', '#00FFFFFF']: # 忽略默认黑白
                cell_css += f"color: {font_color}; "
                
            row_style.append(cell_css)
            
            # 3. 备注
            if cell.comment:
                row_comment.append(f"[{cell.coordinate}]: {cell.comment.text}")
            else:
                row_comment.append(None)
                
        data.append(row_data)
        styles.append(row_style)
        comments.append(row_comment)
        
    df = pd.DataFrame(data, columns=headers)
    style_df = pd.DataFrame(styles, columns=headers)
    
    # 提取所有备注到单独的列表，方便展示
    all_comments = []
    for r_idx, row_c in enumerate(comments):
        row_notes = [c for c in row_c if c]
        if row_notes:
            all_comments.append({"Row": r_idx + 2, "Notes": "; ".join(row_notes)})
            
    comments_df = pd.DataFrame(all_comments)
            
    return df, style_df, comments_df, None

# 主逻辑
file_to_load = uploaded_file if uploaded_file else DEFAULT_FILE_PATH

if file_to_load:
    try:
        if isinstance(file_to_load, str) and not os.path.exists(file_to_load):
             st.warning("⚠️ 默认文件不存在，请上传文件。")
        else:
            with st.spinner('正在解析 Excel 文件及样式...'):
                df, style_df, comments_df, error_msg = load_data_with_styles(file_to_load)
            
            if error_msg:
                st.error(error_msg)
            else:
                # -----------------
                # 2. 筛选与控制区
                # -----------------
                st.sidebar.header("2. 筛选控制")
                
                # 列筛选（隐藏列）
                all_columns = df.columns.tolist()
                default_cols = all_columns[:10] # 默认显示前10列
                selected_cols = st.sidebar.multiselect("选择要显示的列", all_columns, default=all_columns)
                
                # 内容筛选 (示例：筛选包含特定关键字的行)
                search_term = st.sidebar.text_input("全文搜索 (过滤行)", "")
                
                # -----------------
                # 3. 数据处理与展示
                # -----------------
                
                # 应用筛选
                filtered_df = df.copy()
                filtered_style_df = style_df.copy()
                
                if search_term:
                    # 简单全文搜索
                    mask = filtered_df.astype(str).apply(lambda x: x.str.contains(search_term, case=False, na=False)).any(axis=1)
                    filtered_df = filtered_df[mask]
                    filtered_style_df = filtered_style_df[mask]
                
                # 只保留选中的列
                if selected_cols:
                    filtered_df = filtered_df[selected_cols]
                    filtered_style_df = filtered_style_df[selected_cols]
                
                st.subheader(f"数据预览 ({len(filtered_df)} 行)")
                
                # 应用样式
                # 定义样式应用函数
                def style_apply(x):
                    # x 是一个 DataFrame，我们需要返回一个同样形状的 DataFrame，包含 CSS 字符串
                    # 这里我们需要根据原始索引找到对应的 style_df
                    # 注意：pandas styler apply 是按列或按行处理的，或者 applymap 按单元格
                    # 这里最简单的是直接使用对应的 style 矩阵
                    return filtered_style_df
                
                # 使用 Styler
                # 注意：Styler.apply 需要传递一个函数，该函数接收数据并返回样式
                # 这里我们利用索引对齐
                
                styler = filtered_df.style.apply(lambda _: filtered_style_df, axis=None)
                
                # 展示表格
                st.dataframe(styler, use_container_width=True, height=600)
                
                # -----------------
                # 4. 备注展示区
                # -----------------
                if not comments_df.empty:
                    with st.expander("📝 查看所有单元格备注 (点击展开)", expanded=False):
                        st.table(comments_df)
                        
    except Exception as e:
        st.error(f"发生错误: {str(e)}")
else:
    st.info("请上传一个 Excel 文件以开始分析。")
