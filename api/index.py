from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
import io
from typing import List, Dict, Any

app = FastAPI()

# Enable CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify the frontend origin
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_rgb_color(color_obj):
    if not color_obj:
        return None
    if color_obj.type == 'rgb':
        # Some rgb values are 8 chars (AARRGGBB), we need RRGGBB
        rgb = color_obj.rgb
        if len(rgb) == 8:
            return '#' + rgb[2:]
        return '#' + rgb
    # Handle theme colors if needed, skipping for now as it's complex without theme xml
    return None

import re

def process_d240_faults(rows: List[Dict], headers: List[str]) -> List[Dict]:
    """
    Process '故障代码D240' rows:
    1. Sort by Inner Time, then Inner MS.
    2. Merge faults with identical (Device Time, Inner Time, Inner MS).
    3. Merged format: [InnerTime InnerMS] ['Fault1']['Fault2']...
    """
    if not rows:
        return rows
        
    content_col = next((h for h in headers if "内容" in h), None)
    time_col = next((h for h in headers if "时间" in h), None)
    type_col = next((h for h in headers if "类型" in h), None)
    
    if not content_col or not time_col:
        return rows
        
    # Helper to parse inner timestamp
    # Format: [2025-12-08 10:18:04 120ms] ...
    # Regex groups: (Date Time), (MS)
    inner_time_pattern = re.compile(r"\[(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+(\d+)ms\]")
    
    # Helper to extract fault code for sorting
    # Format: '434(安全回路（#29）断开)'
    fault_code_pattern = re.compile(r"^\s*'([A-Za-z0-9]+)")
    
    # Group rows by Device Time
    # Key: str(device_time_value) -> List of row indices
    device_time_groups = {}
    for i, row in enumerate(rows):
        t_val = row["cells"].get(time_col, {}).get("value")
        t_key = str(t_val) if t_val else "UNKNOWN"
        if t_key not in device_time_groups:
            device_time_groups[t_key] = []
        device_time_groups[t_key].append(i)
        
    # Set of indices to remove (merged into others)
    indices_to_remove = set()
    
    # Map of index -> new row data (if modified/merged)
    modified_rows = {}
    
    for t_key, indices in device_time_groups.items():
        # Identify D240 rows in this group
        d240_indices = []
        for idx in indices:
            row = rows[idx]
            # Check type or content for "故障代码D240"
            # User said: "故障代码D240" is likely in type column or content
            type_val = str(row["cells"].get(type_col, {}).get("value", "")) if type_col else ""
            content_val = str(row["cells"].get(content_col, {}).get("value", ""))
            
            if "故障代码D240" in type_val or "故障代码D240" in content_val:
                d240_indices.append(idx)
        
        if not d240_indices:
            continue
            
        # Parse D240 rows
        # List of dicts: {idx, inner_time_str, inner_ms_int, fault_content_str}
        parsed_d240 = []
        for idx in d240_indices:
            content_val = str(rows[idx]["cells"].get(content_col, {}).get("value", ""))
            match = inner_time_pattern.search(content_val)
            if match:
                inner_time = match.group(1)
                inner_ms = int(match.group(2))
                # Extract fault content part: everything after ]
                # e.g. " ['697(...)']"
                fault_part = content_val[match.end():].strip()
            else:
                # Fallback if pattern doesn't match, treat as 0 timestamp?
                # Or just ignore sorting for this one
                inner_time = "0000-00-00 00:00:00"
                inner_ms = 0
                fault_part = content_val
            
            parsed_d240.append({
                "idx": idx,
                "inner_time": inner_time,
                "inner_ms": inner_ms,
                "fault_part": fault_part
            })
            
        # Group by (inner_time, inner_ms) for merging
        merge_groups = {}
        for item in parsed_d240:
            key = (item["inner_time"], item["inner_ms"])
            if key not in merge_groups:
                merge_groups[key] = []
            merge_groups[key].append(item)
            
        # Process each merge group
        # Result: List of new row objects (merged)
        final_d240_rows = []
        
        # We need to process merge groups in order of time key
        sorted_keys = sorted(merge_groups.keys(), key=lambda k: (k[0], k[1]))
        
        for key in sorted_keys:
            items = merge_groups[key]
            
            # Collect all faults
            all_faults = []
            for item in items:
                # fault_part might be "['Code(...)']" or multiple "['A'] ['B']" ?
                # User example: "['697(...)']"
                # We need to extract individual ['...'] blocks or just parse the strings
                # Let's simple find all matches of \['[^']+?\'\]
                # or just use the fault_part string if it's simple.
                # User wants to sort by code.
                
                # Regex to find ['...'] blocks
                # Assuming standard python list repr or user's bracketed format
                # Example: "['434(...)']"
                
                # Let's find all occurrences of ['...']
                # Be careful with nested brackets if any, but fault codes usually simple.
                
                fault_matches = re.findall(r"(\['[^']+'\])", item["fault_part"])
                if not fault_matches:
                    # Maybe it's not bracketed? Just add whole part
                    all_faults.append((item["fault_part"], "0"))
                else:
                    for f_str in fault_matches:
                        # Extract code for sorting
                        # f_str is "['434(...)']"
                        # Clean to "434(...)"
                        inner_str = f_str[2:-2] # remove [' and ']
                        # Extract code
                        code_match = fault_code_pattern.search(inner_str)
                        code = code_match.group(1) if code_match else inner_str
                        all_faults.append((f_str, code))
            
            # Sort faults by code
            # Assuming alphanumeric sort
            all_faults.sort(key=lambda x: x[1])
            
            # Construct merged content
            # Format: [Time MS] ['Fault1']['Fault2']
            inner_time, inner_ms = key
            merged_faults_str = "".join([f[0] for f in all_faults])
            new_content = f"[{inner_time} {inner_ms}ms] {merged_faults_str}"
            
            # We use the FIRST index of this group to hold the result
            # But wait, we are reordering everything within the Device Time group.
            # So we will just generate a list of row contents to place back into the d240_indices slots.
            
            final_d240_rows.append(new_content)
            
        # Now we have sorted, merged contents: final_d240_rows
        # And we have original indices: d240_indices
        
        # We place the new contents into the first N indices of d240_indices
        # And mark the remaining indices for removal
        
        for i in range(len(d240_indices)):
            orig_idx = d240_indices[i]
            if i < len(final_d240_rows):
                # Update this row
                # We need to deep copy the row to avoid mutating original list reference directly until we build result
                # But here we can just prepare a modification
                new_content = final_d240_rows[i]
                
                # Clone row
                old_row = rows[orig_idx]
                new_row = old_row.copy()
                new_row["cells"] = old_row["cells"].copy()
                new_row["cells"][content_col] = old_row["cells"][content_col].copy()
                new_row["cells"][content_col]["value"] = new_content
                
                modified_rows[orig_idx] = new_row
            else:
                # This slot is no longer needed (merged into previous)
                indices_to_remove.add(orig_idx)
                
    # Reconstruct rows list
    new_rows = []
    for i, row in enumerate(rows):
        if i in indices_to_remove:
            continue
        if i in modified_rows:
            new_rows.append(modified_rows[i])
        else:
            new_rows.append(row)
            
    return new_rows

def aggregate_traces(rows: List[Dict], headers: List[str]) -> List[Dict]:
    """
    Aggregate Control Trace (53552 center) and Management Trace (53504 center).
    """
    if not rows:
        return rows

    # Find the index of "信息内容" or similar column
    content_col = next((h for h in headers if "内容" in h), None)
    time_col = next((h for h in headers if "时间" in h), None)
    
    if not content_col:
        return rows

    # Helper to extract ID from content
    def extract_id(text):
        if not text:
            return None
        match = re.search(r'Trace[:：]\s*(\d+)', str(text))
        if match:
            return match.group(1)
        return None

    # Helper to extract timestamp string from content (e.g. [2025...])
    def extract_content_timestamp(text):
        if not text:
            return ""
        match = re.search(r'(\[[^\]]+\])', str(text))
        if match:
            return match.group(1)
        return ""

    processed_rows = []
    skip_indices = set()
    
    # We iterate by index to look ahead/behind
    # Group rows by timestamp first? 
    # Actually, scanning neighbors is safer because they might not be strictly sorted by time if milliseconds differ slightly or are missing.
    # But usually they are adjacent.
    
    # Let's group by timestamp first to make it robust
    # Map: timestamp -> list of (index, row)
    # But "timestamp" column value might differ slightly? 
    # The example shows identical "装置时间": 2025-12-08 10:58:17
    
    rows_by_time = {}
    for i, row in enumerate(rows):
        t_val = row["cells"].get(time_col, {}).get("value")
        if t_val:
            # Convert to string to use as key
            t_key = str(t_val)
            if t_key not in rows_by_time:
                rows_by_time[t_key] = []
            rows_by_time[t_key].append((i, row))
        else:
            # No time, treat as unique group? Or just separate.
            # We'll just append them to processed_rows if they are not skipped
            pass

    # Target sets
    control_target = {'53552', '53553', '53554', '53555', '53556', '53557', '53558'}
    mgmt_target = {'53504', '53505', '53506', '53507', '53508'}
    
    replacements = {} # Map index -> new content string

    # First pass: Identify clusters and mark rows to skip
    for i, row in enumerate(rows):
        # We only look for centers here
        content_cell = row["cells"].get(content_col, {})
        content_text = content_cell.get("value", "")
        trace_id = extract_id(content_text)
        
        if trace_id == '53552':
            # Control Trace Center
            t_val = row["cells"].get(time_col, {}).get("value")
            t_key = str(t_val) if t_val else None
            
            candidates = rows_by_time.get(t_key, [(i, row)])
            
            found_ids = set()
            cluster_indices = []
            
            for c_idx, c_row in candidates:
                # Don't claim rows already claimed by another cluster?
                # But here we are just finding members.
                # If a row is a member of multiple clusters (impossible with current ID sets), we might have issue.
                # Assuming disjoint sets.
                
                c_text = c_row["cells"].get(content_col, {}).get("value", "")
                c_id = extract_id(c_text)
                
                if c_id in control_target:
                    found_ids.add(c_id)
                    cluster_indices.append(c_idx)
            
            # Generate summary
            timestamp_str = extract_content_timestamp(content_text)
            missing = sorted(list(control_target - found_ids))
            
            if not missing:
                summary = f"控制Trace{timestamp_str}（完整）"
            else:
                missing_str = "、".join(missing)
                summary = f"控制Trace{timestamp_str} 缺少{missing_str}数据"
            
            replacements[i] = summary
            
            # Mark members to skip (excluding the center itself, which we will handle via replacements)
            for c_idx in cluster_indices:
                if c_idx != i:
                    skip_indices.add(c_idx)

        elif trace_id == '53504':
            # Management Trace Center
            t_val = row["cells"].get(time_col, {}).get("value")
            t_key = str(t_val) if t_val else None
            
            candidates = rows_by_time.get(t_key, [(i, row)])
            
            found_ids = set()
            cluster_indices = []
            
            for c_idx, c_row in candidates:
                c_text = c_row["cells"].get(content_col, {}).get("value", "")
                c_id = extract_id(c_text)
                
                if c_id in mgmt_target:
                    found_ids.add(c_id)
                    cluster_indices.append(c_idx)
            
            # Generate summary
            timestamp_str = extract_content_timestamp(content_text)
            missing = sorted(list(mgmt_target - found_ids))
            
            if not missing:
                summary = f"管理Trace{timestamp_str}（完整）"
            else:
                missing_str = "、".join(missing)
                summary = f"管理Trace{timestamp_str} 缺少{missing_str}数据"
            
            replacements[i] = summary
            
            for c_idx in cluster_indices:
                if c_idx != i:
                    skip_indices.add(c_idx)

    # Second pass: Build result
    for i, row in enumerate(rows):
        if i in skip_indices:
            continue
            
        if i in replacements:
            new_row = row.copy()
            new_row["cells"] = row["cells"].copy()
            new_row["cells"][content_col] = row["cells"][content_col].copy()
            new_row["cells"][content_col]["value"] = replacements[i]
            processed_rows.append(new_row)
        else:
            processed_rows.append(row)
            
    return processed_rows

import math

def clean_for_json(obj):
    """
    Recursively clean dictionary/list to replace NaN/Infinity with None or string,
    ensuring JSON compatibility for standard JSON parsers (like in browsers).
    """
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    elif isinstance(obj, dict):
        return {k: clean_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_for_json(v) for v in obj]
    return obj

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file.")
    
    content = await file.read()
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        
        # Find the target sheet
        target_sheet_name = None
        for name in wb.sheetnames:
            if "时间线" in name or "Timeline" in name:
                target_sheet_name = name
                break
        
        if not target_sheet_name:
            # Fallback to first sheet if no timeline sheet found
            target_sheet_name = wb.sheetnames[0]
            
        ws = wb[target_sheet_name]
        
        headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]
        
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = {
                "id": row_idx,
                "cells": {}
            }
            
            # Check if row has any content
            has_content = False
            
            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    break
                
                col_name = headers[col_idx]
                val = cell.value
                
                if val is not None:
                    has_content = True
                
                # Extract styles
                bg_color = get_rgb_color(cell.fill.fgColor) if cell.fill else None
                font_color = get_rgb_color(cell.font.color) if cell.font else None
                
                # Extract comment
                comment = cell.comment.text if cell.comment else None
                
                cell_data = {
                    "value": val,
                    "style": {},
                    "comment": comment
                }
                
                if bg_color and bg_color.upper() not in ['#000000', '#FFFFFF', '#00FFFFFF']:
                    cell_data["style"]["backgroundColor"] = bg_color
                
                if font_color and font_color.upper() not in ['#000000', '#00FFFFFF']:
                    cell_data["style"]["color"] = font_color
                
                row_data["cells"][col_name] = cell_data
            
            if has_content:
                rows.append(row_data)
        
        # Aggregate trace data
        rows = aggregate_traces(rows, headers)
        
        # Process D240 faults (sort and merge)
        rows = process_d240_faults(rows, headers)
        
        result = {
            "sheet_name": target_sheet_name,
            "headers": headers,
            "rows": rows
        }
        
        return clean_for_json(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
