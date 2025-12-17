from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
import io
from typing import List, Dict, Any, Union
import os
from openai import OpenAI
from pydantic import BaseModel

app = FastAPI()

# Initialize OpenAI Client (Ensure OPENAI_API_KEY is set in environment)
client = OpenAI(
    api_key=os.environ.get("OPENAI_API_KEY", "sk-placeholder"),
    base_url=os.environ.get("OPENAI_BASE_URL", "https://api.openai.com/v1")
)

class AnalysisRequest(BaseModel):
    rows: List[Dict]
    context: str = ""

# Enable CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify the frontend origin
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_rgb_color(color_obj):
    if not color_obj:
        return None
    if color_obj.type == 'rgb':
        # Some rgb values are 8 chars (AARRGGBB), we need RRGGBB
        rgb = color_obj.rgb
        if len(rgb) == 8:
            return '#' + rgb[2:]
        return '#' + rgb
    # Handle theme colors if needed, skipping for now as it's complex without theme xml
    return None

import re
import json
from datetime import datetime

class TimelinePreprocessor:
    """
    Implements domain-specific logic to clean, filter, and enrich timeline data
    before sending it to the LLM.
    """
    
    # Configuration for Tagging System
    TAG_CONFIG = {
        "non_critical": {
            "keywords": [
                "391(管理综合故障)",
                "7C8(前门门机警告类综合故障)",
                "00305025(前门关门轻故障)"
            ],
            "regex": [
                # r"故障代码.*\(无关\)" # Example regex
            ]
        },
        "delayed_upload": {
            "threshold_minutes": 3
        }
    }

    # Configuration for Global Attribute Extraction
    # Format: { AttributeName: [ { keys: [k1, k2], value_map: {val_in_json: normalized_val} } ] }
    GLOBAL_ATTR_CONFIG = {
        "合同号": [
             {
                 "keys": ["产品合同号梯号", "Contract No.", "Device ID"],
                 "value_map": None
             }
        ],
        "控制同步层": [
            {
                "keys": ["控制同步层"], # Try these keys in order
                "value_map": None, # Direct value
                "transform": lambda x: int(x) - 1 if str(x).isdigit() else x # Transform +1 for floor
            }
        ],
        "41DG信号": [
            # Case 1: 故障诊断履历
            {
                "keys": ["41DG信号"],
                "value_map": {
                    "闭合": "闭合",
                    "断开": "断开"
                }
            },
            # Case 2: 运行模式 / 检修开关履历
            {
                "keys": ["门锁状态（41DG）"],
                "value_map": {
                    "门锁断开(41DG_OFF)": "断开",
                    "门锁闭合(41DG_ON)": "闭合", # Assuming opposite exists
                    "闭合": "闭合"
                }
            }
        ]
    }

    def __init__(self):
        self.last_fault_time = 0
        self.last_fault_code = None

    def _is_purple_color(self, hex_color: str) -> bool:
        if not hex_color:
            return False
        hex_color = hex_color.upper().replace('#', '')
        if len(hex_color) != 6:
            return False
        try:
            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
            return r > 100 and b > 100 and g < 100
        except:
            return False

    def _extract_json_from_comment(self, comment: str) -> Dict:
        """
        Robustly extract JSON-like structure from comment string.
        Handles: Note: {...} or just {...} or Python dict string representation.
        """
        if not comment:
            return {}
        
        # Try to find { ... } block
        match = re.search(r'(\{.*\})', comment, re.DOTALL)
        if not match:
            return {}
        
        json_str = match.group(1)
        try:
            # Try standard JSON
            return json.loads(json_str)
        except:
            try:
                # Try replacing single quotes with double quotes (Python dict str)
                # Handle boolean values
                fixed_str = json_str.replace("'", '"')\
                                    .replace("True", "true")\
                                    .replace("False", "false")\
                                    .replace("None", "null")
                return json.loads(fixed_str)
            except:
                return {}

    def _get_value_from_cells(self, row: Dict, key_substr: str) -> str:
        """Helper to find value in cells where column name contains key_substr"""
        for col_name, cell_data in row.get("cells", {}).items():
            if key_substr in col_name:
                return str(cell_data.get("value", ""))
        return ""

    def _get_comment_from_cells(self, row: Dict) -> str:
        """Combine comments from all cells"""
        comments = []
        for cell_data in row.get("cells", {}).values():
            c = cell_data.get("comment")
            if c:
                comments.append(c)
        return "\n".join(comments)

    def _parse_timestamp(self, time_str: str) -> datetime:
        """Parse timestamp string to datetime object."""
        if not time_str:
            return None
        # Try common formats
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%H:%M:%S" # If no date, might need context, but for diff check ok?
        ]
        for fmt in formats:
            try:
                # If only time, assume today or handle gracefully? 
                # For 10 min diff check, we need full datetime usually.
                # If input has only time, we might fail comparison if date differs.
                # Assuming input string has Date if "装置时间" has Date.
                return datetime.strptime(time_str, fmt)
            except ValueError:
                continue
        return None

    def _extract_attributes(self, row: Dict) -> Dict[str, Any]:
        """
        Extract global attributes from a single row.
        Priority:
        1. Extract from Note JSON (comments)
        2. Extract from Cell Values (columns)
        3. Special calculations (e.g., Delay Duration)
        """
        extracted = {}
        # Extract Note JSON
        note_str = self._get_comment_from_cells(row)
        note_json = self._extract_json_from_comment(note_str)
        
        for attr_name, rules in self.GLOBAL_ATTR_CONFIG.items():
            found_val = None
            
            for rule in rules:
                # Strategy 1: Try to find in Note JSON
                if note_json:
                    for k in rule['keys']:
                        if k in note_json:
                            found_val = note_json[k]
                            break
                
                # Strategy 2: If not found, try to find in Cell Values (Column Names)
                if found_val is None:
                    for k in rule['keys']:
                        # Check if any column name contains this key
                        val_from_col = self._get_value_from_cells(row, k)
                        if val_from_col:
                            found_val = val_from_col
                            break
                
                if found_val is not None:
                    # Apply Value Mapping if exists
                    if rule.get('value_map'):
                        if found_val in rule['value_map']:
                            found_val = rule['value_map'][found_val]
                        elif str(found_val) in rule['value_map']:
                            found_val = rule['value_map'][str(found_val)]
                        else:
                            # If map exists but value not found, keep raw? 
                            # Or strictly map? 
                            # Assuming keep raw unless explicit strict mode.
                            pass
                            
                    # Apply Transformation if exists
                    if rule.get('transform'):
                        try:
                            found_val = rule['transform'](found_val)
                        except:
                            pass
                    
                    break # Found value for this rule
            
            if found_val is not None:
                extracted[attr_name] = found_val
        
        # Strategy 3: Special Calculation - Delay Duration
        # Calculate delay regardless of threshold for attribute display
        content_val = self._get_value_from_cells(row, "内容") or self._get_value_from_cells(row, "Content")
        time_str = self._get_value_from_cells(row, "时间") or self._get_value_from_cells(row, "Time")
        
        if content_val and time_str:
            timestamp_pattern = re.compile(r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})")
            content_ts_match = timestamp_pattern.search(content_val)
            if content_ts_match:
                content_ts_str = content_ts_match.group(1)
                try:
                    row_dt = self._parse_timestamp(time_str)
                    content_dt = self._parse_timestamp(content_ts_str)
                    
                    if row_dt and content_dt:
                        diff = (row_dt - content_dt).total_seconds() / 60
                        # Only record if positive delay? Or all? 
                        # Usually "Delay" implies positive. 
                        # If diff > 0, it means uploaded later than event time.
                        if diff > 0:
                            extracted["延时时长"] = f"{int(diff)}m"
                except:
                    pass

        return extracted

    def _get_tags(self, row: Dict) -> List[str]:
        """
        Get tags for a single row based on rules.
        """
        tags = []
        content_val = self._get_value_from_cells(row, "内容") or self._get_value_from_cells(row, "Content")
        time_str = self._get_value_from_cells(row, "时间") or self._get_value_from_cells(row, "Time")
        
        if not content_val:
            return tags
            
        # A. Non-Critical Tagging
        is_non_critical = False
        # Check keywords
        for kw in self.TAG_CONFIG["non_critical"]["keywords"]:
            if kw in content_val:
                is_non_critical = True
                break
        # Check regex
        if not is_non_critical:
            for rx in self.TAG_CONFIG["non_critical"]["regex"]:
                if re.search(rx, content_val):
                    is_non_critical = True
                    break
        
        if is_non_critical:
            tags.append("【ℹ️非关键】")

        # B. Delayed Upload Tagging
        timestamp_pattern = re.compile(r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})")
        content_ts_match = timestamp_pattern.search(content_val)
        if content_ts_match and time_str:
            content_ts_str = content_ts_match.group(1)
            try:
                row_dt = self._parse_timestamp(time_str)
                content_dt = self._parse_timestamp(content_ts_str)
                
                if row_dt and content_dt:
                    diff = (row_dt - content_dt).total_seconds() / 60
                    if diff > self.TAG_CONFIG["delayed_upload"]["threshold_minutes"]:
                        tags.append(f"【⏳延时上传:{int(diff)}分】")
            except:
                pass

        # C. Human Operation Tagging
        is_human = False
        cells = row.get("cells", {})
        if isinstance(cells, dict):
            for cell in cells.values():
                if isinstance(cell, dict):
                    style = cell.get("style", {})
                    if isinstance(style, dict):
                        if self._is_purple_color(style.get("backgroundColor")):
                            is_human = True
                            break
        if "检修" in content_val or "机修工单" in content_val:
            is_human = True
        if is_human:
            tags.append("【⚠️现场人工操作】")

        # D. Work Order Tagging
        if "工单" in content_val:
            tags.append("【🚨高优先级-工单】")
            
        return tags

    def process(self, rows: List[Dict], return_logs: bool = False) -> Union[str, Dict]:
        """
        Main entry point. 
        If return_logs is True, returns a dict with 'text' and 'logs'.
        Otherwise returns just the enriched text string.
        """
        enriched_lines = []
        debug_logs = {
            "ignored_rows": [],      # List of {id, time, content, reason}
            "delayed_rows": [],      # List of {id, time, content, delay_min}
            "attribute_rows": []     # List of {id, time, content, extracted_attrs}
        }
        
        # We need "装置时间" (Device Time) to check for delay.
        # Assuming "时间" column is the upload time/log time, and there might be an inner "Device Time"?
        # Or user means: Content has "Time" and Row has "Time".
        # User said: "只要信息内容中有时间信息，且该时间比装置时间早超过10分钟".
        # Let's assume Row Time = Device Time (or Upload Time), and Content might contain another timestamp.
        # Actually usually: Row Time = Log Time (Device Time). 
        # Wait, if "Upload Time" is later than "Device Time"? 
        # Let's look for timestamp in Content.
        
        timestamp_pattern = re.compile(r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})")

        for row in rows:
            # 1. Get Basic Info
            row_id = row.get("id")
            time_str = self._get_value_from_cells(row, "时间") or self._get_value_from_cells(row, "Time")
            content_val = self._get_value_from_cells(row, "内容") or self._get_value_from_cells(row, "Content")
            
            # Skip if empty content
            if not content_val:
                continue

            tags = []

            # === 2. Rule-Based Tagging ===
            
            # A. Non-Critical Tagging
            is_non_critical = False
            # Check keywords
            for kw in self.TAG_CONFIG["non_critical"]["keywords"]:
                if kw in content_val:
                    is_non_critical = True
                    break
            # Check regex
            if not is_non_critical:
                for rx in self.TAG_CONFIG["non_critical"]["regex"]:
                    if re.search(rx, content_val):
                        is_non_critical = True
                        break
            
            if is_non_critical:
                tags.append("【ℹ️非关键】")
                if return_logs:
                    debug_logs["ignored_rows"].append({
                        "id": row_id,
                        "time": time_str,
                        "content": content_val,
                        "reason": "Matched non-critical keyword/regex"
                    })

            # B. Delayed Upload Tagging
            # "信息内容中有时间信息" -> Look for timestamp in content_val
            content_ts_match = timestamp_pattern.search(content_val)
            if content_ts_match and time_str:
                content_ts_str = content_ts_match.group(1)
                try:
                    # Row time (Device Time / Log Time)
                    row_dt = self._parse_timestamp(time_str)
                    # Content time (Event Time)
                    content_dt = self._parse_timestamp(content_ts_str)
                    
                    if row_dt and content_dt:
                        # Calculate difference in minutes
                        diff = (row_dt - content_dt).total_seconds() / 60
                        if diff > self.TAG_CONFIG["delayed_upload"]["threshold_minutes"]:
                            tags.append(f"【⏳延时上传:{int(diff)}分】")
                            if return_logs:
                                debug_logs["delayed_rows"].append({
                                    "id": row_id,
                                    "time": time_str,
                                    "content": content_val,
                                    "delay_min": int(diff)
                                })
                except:
                    pass # Date parsing failed, skip check

            # C. Human Operation Tagging (Purple)
            is_human = False
            cells = row.get("cells", {})
            if isinstance(cells, dict):
                for cell in cells.values():
                    if isinstance(cell, dict):
                        style = cell.get("style", {})
                        if isinstance(style, dict):
                            if self._is_purple_color(style.get("backgroundColor")):
                                is_human = True
                                break
            if "检修" in content_val or "机修工单" in content_val:
                is_human = True
            if is_human:
                tags.append("【⚠️现场人工操作】")

            # D. Work Order Tagging
            if "工单" in content_val:
                tags.append("【🚨高优先级-工单】")
            
            # === 3. Global Attribute Extraction ===
            # Use the helper method
            extracted_attrs_dict = self._extract_attributes(row)
            extracted_signals = [f"{k}={v}" for k, v in extracted_attrs_dict.items()]

            if extracted_signals and return_logs:
                debug_logs["attribute_rows"].append({
                    "id": row_id,
                    "time": time_str,
                    "content": content_val,
                    "extracted_attrs": extracted_signals
                })

            # === 4. Construct Final Line ===
            
            # Filter out non-critical lines IF we want to hide them from LLM to save tokens
            # But user said "打标", maybe LLM needs to know it's non-critical but still see it?
            # Or just hide it. Let's hide it if it's explicitly "Non-Critical" to reduce noise.
            # However, for "Delayed Upload", we want to show it.
            
            if "【ℹ️非关键】" in tags:
                # We can choose to skip adding this line to context
                # return or continue
                continue 

            line = f"[{time_str}]"
            if tags:
                line += " " + " ".join(tags)
            line += f" {content_val}"
            
            if extracted_signals:
                line += "\n   >> 全局属性: " + ", ".join(extracted_signals)
                
            enriched_lines.append(line)
            
        text_result = "\n".join(enriched_lines)
        
        if return_logs:
            return {"text": text_result, "logs": debug_logs}
        return text_result

@app.post("/api/debug/preview_rules")
async def preview_rules(request: AnalysisRequest):
    """
    Debug endpoint to preview how rules are applied to the data.
    Returns detailed logs of ignored rows, delayed uploads, and extracted attributes.
    """
    try:
        preprocessor = TimelinePreprocessor()
        result = preprocessor.process(request.rows, return_logs=True)
        return result["logs"]
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/analyze")
async def analyze_events(request: AnalysisRequest):
    """
    Analyze a list of events using LLM with domain knowledge injection.
    """
    try:
        # 1. Preprocess rows using Domain Logic
        preprocessor = TimelinePreprocessor()
        data_context = preprocessor.process(request.rows)
        
        if not data_context:
            return {"analysis": "根据预设规则，所选数据中没有发现高价值信息（可能已被过滤）。"}

        # 2. Construct Prompt
        system_prompt = """
你是一个电梯工业数据分析专家。请基于提供的时间线事件数据进行解读。

【分析原则】
1. **高优先级**：工单 > 故障发生(有效) > 故障恢复。这些通常意味着发生了停梯或严重问题。
2. **人工操作**：标记为【⚠️现场人工操作】的时间段，代表维保人员在现场。在此期间产生的故障可能是调试过程，需结合上下文区分。
3. **关键信号**：
   - 关注 '关键信号' 行的数据。
   - 安全回路(Safety Circuit)断开通常是故障根源。
   - 门锁回路(Door Lock)断开会导致电梯急停。
4. **忽略项**：已被标记为【⬇️已降权】或未出现在列表中的警告类信息请忽略。

【输出要求】
1. **结论先行**：第一句话直接告诉用户发生了什么（例如：“电梯在4楼因安全回路断开导致急停，随后维保人员到场检修”）。
2. **证据链**：列出支持你结论的关键事件和时间点。
3. **排版**：使用 Markdown，重点加粗。
"""
        
        user_prompt = f"""
请分析以下事件数据：

{data_context}

用户补充背景：
{request.context}
"""

        # 3. Call LLM
        response = client.chat.completions.create(
            model="gpt-3.5-turbo", 
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.3
        )
        
        return {"analysis": response.choices[0].message.content}
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

def process_d240_faults(rows: List[Dict], headers: List[str]) -> List[Dict]:
    """
    Process '故障代码D240' rows:
    1. Group by Device ID (Contract No.) to ensure isolation.
    2. Sort by Inner Time, then Inner MS.
    3. Merge faults with identical (Device Time, Inner Time, Inner MS).
    4. Merged format: [InnerTime InnerMS] ['Fault1']['Fault2']...
    """
    if not rows:
        return rows
        
    content_col = next((h for h in headers if "内容" in h), None)
    time_col = next((h for h in headers if "时间" in h), None)
    type_col = next((h for h in headers if "类型" in h), None)
    # Identify device/contract column (usually 2nd column)
    device_col = headers[1] if len(headers) > 1 else None
    
    if not content_col or not time_col:
        return rows
        
    # Helper to parse inner timestamp
    # Format: [2025-12-08 10:18:04 120ms] ...
    # Regex groups: (Date Time), (MS)
    inner_time_pattern = re.compile(r"\[(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+(\d+)ms\]")
    
    # Helper to extract fault code for sorting
    # Format: '434(安全回路（#29）断开)'
    fault_code_pattern = re.compile(r"^\s*'([A-Za-z0-9]+)")
    
    # --- Group by Device ID first ---
    # device_id -> list of indices
    device_groups = {}
    for i, row in enumerate(rows):
        d_val = "UNKNOWN"
        if device_col:
            d_val = str(row["cells"].get(device_col, {}).get("value", "UNKNOWN"))
        
        if d_val not in device_groups:
            device_groups[d_val] = []
        device_groups[d_val].append(i)
    
    # Set of indices to remove (merged into others)
    indices_to_remove = set()
    # Map of index -> new row data (if modified/merged)
    modified_rows = {}
    
    # Process each device group independently
    for d_val, d_indices in device_groups.items():
        
        # Sub-group by Device Time within this device
        # Key: str(device_time_value) -> List of row indices
        device_time_groups = {}
        for idx in d_indices:
            row = rows[idx]
            t_val = row["cells"].get(time_col, {}).get("value")
            t_key = str(t_val) if t_val else "UNKNOWN"
            if t_key not in device_time_groups:
                device_time_groups[t_key] = []
            device_time_groups[t_key].append(idx)
            
        for t_key, indices in device_time_groups.items():
            # Identify D240 rows in this group
            d240_indices = []
            for idx in indices:
                row = rows[idx]
                type_val = str(row["cells"].get(type_col, {}).get("value", "")) if type_col else ""
                content_val = str(row["cells"].get(content_col, {}).get("value", ""))
                
                if "故障代码D240" in type_val or "故障代码D240" in content_val:
                    d240_indices.append(idx)
            
            if not d240_indices:
                continue
                
            # Parse D240 rows
            parsed_d240 = []
            for idx in d240_indices:
                content_val = str(rows[idx]["cells"].get(content_col, {}).get("value", ""))
                match = inner_time_pattern.search(content_val)
                if match:
                    inner_time = match.group(1)
                    inner_ms = int(match.group(2))
                    fault_part = content_val[match.end():].strip()
                else:
                    inner_time = "0000-00-00 00:00:00"
                    inner_ms = 0
                    fault_part = content_val
                
                parsed_d240.append({
                    "idx": idx,
                    "inner_time": inner_time,
                    "inner_ms": inner_ms,
                    "fault_part": fault_part
                })
                
            # Group by (inner_time, inner_ms) for merging
            merge_groups = {}
            for item in parsed_d240:
                key = (item["inner_time"], item["inner_ms"])
                if key not in merge_groups:
                    merge_groups[key] = []
                merge_groups[key].append(item)
                
            # Process each merge group
            final_d240_rows = []
            sorted_keys = sorted(merge_groups.keys(), key=lambda k: (k[0], k[1]))
            
            for key in sorted_keys:
                items = merge_groups[key]
                all_faults = []
                for item in items:
                    fault_matches = re.findall(r"(\['[^']+'\])", item["fault_part"])
                    if not fault_matches:
                        all_faults.append((item["fault_part"], "0"))
                    else:
                        for f_str in fault_matches:
                            inner_str = f_str[2:-2]
                            code_match = fault_code_pattern.search(inner_str)
                            code = code_match.group(1) if code_match else inner_str
                            all_faults.append((f_str, code))
                
                all_faults.sort(key=lambda x: x[1])
                inner_time, inner_ms = key
                merged_faults_str = "".join([f[0] for f in all_faults])
                new_content = f"[{inner_time} {inner_ms}ms] {merged_faults_str}"
                final_d240_rows.append(new_content)
                
            # Update rows
            for i in range(len(d240_indices)):
                orig_idx = d240_indices[i]
                if i < len(final_d240_rows):
                    new_content = final_d240_rows[i]
                    old_row = rows[orig_idx]
                    new_row = old_row.copy()
                    new_row["cells"] = old_row["cells"].copy()
                    new_row["cells"][content_col] = old_row["cells"][content_col].copy()
                    new_row["cells"][content_col]["value"] = new_content
                    modified_rows[orig_idx] = new_row
                else:
                    indices_to_remove.add(orig_idx)
                
    # Reconstruct rows list
    new_rows = []
    for i, row in enumerate(rows):
        if i in indices_to_remove:
            continue
        if i in modified_rows:
            new_rows.append(modified_rows[i])
        else:
            new_rows.append(row)
            
    return new_rows

def aggregate_traces(rows: List[Dict], headers: List[str]) -> List[Dict]:
    """
    Aggregate Control Trace (53552 center) and Management Trace (53504 center).
    Ensures aggregation is isolated by Device ID (Contract No.).
    """
    if not rows:
        return rows

    content_col = next((h for h in headers if "内容" in h), None)
    time_col = next((h for h in headers if "时间" in h), None)
    # Identify device/contract column
    device_col = headers[1] if len(headers) > 1 else None
    
    if not content_col:
        return rows

    # Helper to extract ID from content
    def extract_id(text):
        if not text:
            return None
        match = re.search(r'Trace[:：]\s*(\d+)', str(text))
        if match:
            return match.group(1)
        return None

    # Helper to extract timestamp string from content (e.g. [2025...])
    def extract_content_timestamp(text):
        if not text:
            return ""
        match = re.search(r'(\[[^\]]+\])', str(text))
        if match:
            return match.group(1)
        return ""

    processed_rows = []
    skip_indices = set()
    replacements = {} # Map index -> new content string
    
    # --- Group by Device ID first ---
    # device_id -> list of indices
    device_groups = {}
    for i, row in enumerate(rows):
        d_val = "UNKNOWN"
        if device_col:
            d_val = str(row["cells"].get(device_col, {}).get("value", "UNKNOWN"))
        
        if d_val not in device_groups:
            device_groups[d_val] = []
        device_groups[d_val].append(i)

    # Helper to parse timestamp
    def parse_time(t_str):
        if not t_str: return None
        # Try common formats
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%H:%M:%S"]:
            try:
                return datetime.strptime(str(t_str), fmt)
            except ValueError:
                continue
        return None

    # Process each device group independently
    for d_val, d_indices in device_groups.items():
        
        # Target sets
        control_target = {'53552', '53553', '53554', '53555', '53556', '53557', '53558'}
        mgmt_target = {'53504', '53505', '53506', '53507', '53508'}
        
        # Pre-parse timestamps for all rows in this device to optimize lookups
        # Map: idx -> datetime object
        row_times = {}
        for idx in d_indices:
            t_val = rows[idx]["cells"].get(time_col, {}).get("value")
            dt = parse_time(t_val)
            if dt:
                row_times[idx] = dt

        # Iterate over indices in this device group
        for i, idx in enumerate(d_indices):
            row = rows[idx]
            content_cell = row["cells"].get(content_col, {})
            content_text = content_cell.get("value", "")
            trace_id = extract_id(content_text)
            
            if trace_id == '53552': # Control Trace Center
                center_dt = row_times.get(idx)
                
                candidates = []
                if center_dt:
                    # Time-based window: [-10s, +20s]
                    # We scan nearby rows. Since rows are roughly sorted, we can optimize, 
                    # but simple window scan around index is safer if sorting isn't perfect.
                    # Let's scan a reasonable row window (e.g. +/- 50 rows) and check time.
                    search_start = max(0, i - 50)
                    search_end = min(len(d_indices), i + 50)
                    
                    for k in range(search_start, search_end):
                        c_idx = d_indices[k]
                        c_dt = row_times.get(c_idx)
                        if c_dt:
                            # Handle date rollover if formats lack date? 
                            # Assuming formats have date or consistent relative time.
                            # If only time is present, date diff might be huge if crossing midnight, but let's assume simple diff.
                            try:
                                diff = (c_dt - center_dt).total_seconds()
                                if -10 <= diff <= 20:
                                    candidates.append((c_idx, rows[c_idx]))
                            except:
                                pass
                else:
                    # Fallback to row count if no time
                    start_window = max(0, i - 20)
                    end_window = min(len(d_indices), i + 21)
                    for k in range(start_window, end_window):
                        c_idx = d_indices[k]
                        candidates.append((c_idx, rows[c_idx]))
                
                found_ids = set()
                cluster_indices = []
                for c_idx, c_row in candidates:
                    # Skip if this row is already claimed by another cluster (check skip_indices?)
                    # But skip_indices is populated as we go.
                    # If we look BACK, we might pick up a row that was already processed?
                    # The center 53552 is unique for a trace event. 
                    # The parts (53553...) should only belong to one center.
                    # If we have multiple centers close by, we might have ambiguity.
                    # But typically traces are sparse.
                    
                    c_text = c_row["cells"].get(content_col, {}).get("value", "")
                    c_id = extract_id(c_text)
                    if c_id in control_target:
                        # Ensure we don't pick up another center (53552) as a member?
                        # control_target includes 53552.
                        # We are looking for members. 
                        # If c_id is 53552 and c_idx != idx, it's another center. 
                        # We should PROBABLY not merge another center into this one unless it's a duplicate?
                        # But Trace: 53552 IS the center. 
                        # We just want to find unique IDs.
                        
                        # If c_idx is another center (53552) and c_idx != idx, we should ignore it here?
                        # Actually, if we have two 53552s close by, they are likely distinct events.
                        # We should probably only aggregate non-center parts, or parts that haven't been claimed.
                        # For simplicity, let's assume parts are unique in the window.
                        
                        found_ids.add(c_id)
                        cluster_indices.append(c_idx)
                
                timestamp_str = extract_content_timestamp(content_text)
                missing = sorted(list(control_target - found_ids))
                if not missing:
                    summary = f"控制Trace{timestamp_str}（完整）"
                else:
                    missing_str = "、".join(missing)
                    summary = f"控制Trace{timestamp_str} 缺少{missing_str}数据"
                
                replacements[idx] = summary
                for c_idx in cluster_indices:
                    if c_idx != idx:
                        skip_indices.add(c_idx)

            elif trace_id == '53504': # Management Trace Center
                center_dt = row_times.get(idx)
                
                candidates = []
                if center_dt:
                    # Time-based window: [-10s, +20s]
                    search_start = max(0, i - 50)
                    search_end = min(len(d_indices), i + 50)
                    
                    for k in range(search_start, search_end):
                        c_idx = d_indices[k]
                        c_dt = row_times.get(c_idx)
                        if c_dt:
                            try:
                                diff = (c_dt - center_dt).total_seconds()
                                if -10 <= diff <= 20:
                                    candidates.append((c_idx, rows[c_idx]))
                            except:
                                pass
                else:
                    # Fallback
                    start_window = max(0, i - 20)
                    end_window = min(len(d_indices), i + 21)
                    for k in range(start_window, end_window):
                        c_idx = d_indices[k]
                        candidates.append((c_idx, rows[c_idx]))
                
                found_ids = set()
                cluster_indices = []
                for c_idx, c_row in candidates:
                    c_text = c_row["cells"].get(content_col, {}).get("value", "")
                    c_id = extract_id(c_text)
                    if c_id in mgmt_target:
                        found_ids.add(c_id)
                        cluster_indices.append(c_idx)
                
                timestamp_str = extract_content_timestamp(content_text)
                missing = sorted(list(mgmt_target - found_ids))
                if not missing:
                    summary = f"管理Trace{timestamp_str}（完整）"
                else:
                    missing_str = "、".join(missing)
                    summary = f"管理Trace{timestamp_str} 缺少{missing_str}数据"
                
                replacements[idx] = summary
                for c_idx in cluster_indices:
                    if c_idx != idx:
                        skip_indices.add(c_idx)

    # Second pass: Build result
    for i, row in enumerate(rows):
        if i in skip_indices:
            continue
            
        if i in replacements:
            new_row = row.copy()
            new_row["cells"] = row["cells"].copy()
            new_row["cells"][content_col] = row["cells"][content_col].copy()
            new_row["cells"][content_col]["value"] = replacements[i]
            processed_rows.append(new_row)
        else:
            processed_rows.append(row)
            
    return processed_rows

import math

def clean_for_json(obj):
    """
    Recursively clean dictionary/list to replace NaN/Infinity with None or string,
    ensuring JSON compatibility for standard JSON parsers (like in browsers).
    """
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    elif isinstance(obj, dict):
        return {k: clean_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_for_json(v) for v in obj]
    return obj

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file.")
    
    content = await file.read()
    
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        
        # Find the target sheet
        target_sheet_name = None
        for name in wb.sheetnames:
            if "时间线" in name or "Timeline" in name:
                target_sheet_name = name
                break
        
        if not target_sheet_name:
            # Fallback to first sheet if no timeline sheet found
            target_sheet_name = wb.sheetnames[0]
            
        ws = wb[target_sheet_name]
        
        headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]
        
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            row_data = {
                "id": row_idx,
                "cells": {}
            }
            
            # Check if row has any content
            has_content = False
            
            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    break
                
                col_name = headers[col_idx]
                val = cell.value
                
                if val is not None:
                    has_content = True
                
                # Extract styles
                bg_color = get_rgb_color(cell.fill.fgColor) if cell.fill else None
                font_color = get_rgb_color(cell.font.color) if cell.font else None
                
                # Extract comment
                comment = cell.comment.text if cell.comment else None
                
                cell_data = {
                    "value": val,
                    "style": {},
                    "comment": comment
                }
                
                if bg_color and bg_color.upper() not in ['#000000', '#FFFFFF', '#00FFFFFF']:
                    cell_data["style"]["backgroundColor"] = bg_color
                
                if font_color and font_color.upper() not in ['#000000', '#00FFFFFF']:
                    cell_data["style"]["color"] = font_color
                
                row_data["cells"][col_name] = cell_data
            
            if has_content:
                rows.append(row_data)
        
        # Aggregate trace data
        rows = aggregate_traces(rows, headers)
        
        # Process D240 faults (sort and merge)
        rows = process_d240_faults(rows, headers)
        
        # Enrich with global attributes and tags
        preprocessor = TimelinePreprocessor()
        for row in rows:
            row["global_attributes"] = preprocessor._extract_attributes(row)
            row["tags"] = preprocessor._get_tags(row)

        result = {
            "sheet_name": target_sheet_name,
            "headers": headers,
            "rows": rows,
            "server_version": "1.1.0" # Version bump to verify deployment
        }
        
        return clean_for_json(result)

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
